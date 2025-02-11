import os
import random
import pandas as pd
import sqlalchemy as db

from openpyxl import Workbook, load_workbook


def append_dict_to_xlsx(file_path, data_dict):
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(list(data_dict.keys()))
        wb.save(file_path)

    wb = load_workbook(file_path)
    ws = wb.active
    num_rows = len(next(iter(data_dict.values())))
    for i in range(num_rows):
        ws.append([data_dict[key][i] for key in data_dict])

    wb.save(file_path)


data = {"Name": [], "Age": [], "City": []}
names_base = ["Ivan", "Oleg", "Gennadiy", "Oksana", "Olga", "Elena"]
cities_base = ["Moscow", "Omsk", "Saratov"]

for i in range(10000):
    data["Name"].append(random.choice(names_base))
    data["City"].append(random.choice(cities_base))
    data["Age"].append(random.randint(20, 30))


def convert_xlsx_to_csv(xlsx_file, csv_file):
    df = pd.read_excel(xlsx_file)
    df.to_csv(csv_file, index=False)


append_dict_to_xlsx("data.xlsx", data)
convert_xlsx_to_csv("data.xlsx", "data.csv")
